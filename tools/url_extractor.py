import os
import sys
import subprocess
import csv
from io import StringIO
from typing import List, Dict, Tuple, Optional
import openpyxl
import time
import threading
import concurrent.futures as cf
import requests
import json
import re
import dns.resolver
import gzip
import tempfile

def read_whitelist(xlsx_path: str) -> List[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        if "whitelisted domains" in headers:
            col_idx = headers.index("whitelisted domains") + 1
        else:
            col_idx = 4
    except Exception:
        col_idx = 4
    domains = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= col_idx:
            cell = row[col_idx - 1]
            if isinstance(cell, str):
                d = cell.strip().lower()
                if d and "." in d:
                    domains.append(d)
    wb.close()
    seen = set()
    out = []
    for d in domains:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return out

def read_keywords_map(xlsx_path: str) -> Dict[str, List[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        d_idx = headers.index("whitelisted domains") + 1 if "whitelisted domains" in headers else 4
    except Exception:
        d_idx = 4
    try:
        o_idx = headers.index("organisation name") + 1 if "organisation name" in headers else None
    except Exception:
        o_idx = None
    m: Dict[str, List[str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        dom = None
        org = None
        if len(row) >= d_idx:
            cell = row[d_idx - 1]
            if isinstance(cell, str):
                dom = cell.strip().lower()
        if o_idx and len(row) >= o_idx:
            oc = row[o_idx - 1]
            if isinstance(oc, str):
                org = oc.strip().lower()
        if dom and "." in dom:
            parts = dom.split(".")
            sld = parts[-2] if len(parts) >= 2 else parts[0]
            kws = [sld]
            if org:
                clean = re.sub(r"[^a-z0-9 ]+", " ", org)
                words = [w for w in clean.split() if len(w) >= 3]
                if words:
                    kws.append("".join(words))
                    kws.extend(words)
            uniq = []
            seen = set()
            for k in kws:
                if k and k not in seen:
                    seen.add(k)
                    uniq.append(k)
            m[dom] = uniq
    wb.close()
    return m

def read_extra_keywords(path: str) -> List[str]:
    if not os.path.isfile(path):
        return []
    out = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            w = line.strip().lower()
            if w:
                out.append(w)
    seen = set()
    uniq = []
    for w in out:
        if w not in seen:
            seen.add(w)
            uniq.append(w)
    return uniq

def run_dnstwist(domain: str, tlds_path: Optional[str], dict_words: Optional[List[str]] = None) -> Tuple[List[str], List[Dict[str, str]]]:
    cmd = ["dnstwist", "--registered"]
    if tlds_path:
        cmd += ["--tld", tlds_path]
    tmp_path = None
    if dict_words:
        try:
            tf = tempfile.NamedTemporaryFile("w", delete=False, encoding="utf-8")
            for w in dict_words:
                tf.write(w + "\n")
            tf.flush()
            tmp_path = tf.name
            tf.close()
            cmd += ["--dictionary", tmp_path]
        except Exception:
            tmp_path = None
    cmd += ["--format", "csv", domain]
    p = subprocess.run(cmd, capture_output=True, text=True)
    if tmp_path:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass
    if p.returncode != 0 or not p.stdout.strip():
        return [], []
    sio = StringIO(p.stdout)
    reader = csv.reader(sio)
    rows = list(reader)
    if not rows:
        return [], []
    header = rows[0]
    data = []
    for r in rows[1:]:
        item = {}
        for i, k in enumerate(header):
            item[str(k)] = r[i] if i < len(r) else ""
        data.append(item)
    return [str(h) for h in header], data

PLATFORM_BASES = {
    "ngrok": "ngrok.io",
    "vercel": "vercel.app",
    "netlify": "netlify.app",
    "pages": "pages.dev",
    "github": "github.io",
    "gitlab": "gitlab.io",
    "rtdocs": "readthedocs.io",
    "azurestatic": "azurestaticapps.net",
    "azureweb": "azurewebsites.net",
    "azureblob": "web.core.windows.net",
    "appspot": "appspot.com",
    "firebase": "firebaseapp.com",
    "webapp": "web.app",
    "googleusercontent": "googleusercontent.com",
    "googlesites": "sites.google.com",
    "heroku": "herokuapp.com",
    "render": "onrender.com",
    "fly": "fly.dev",
    "railway": "railway.app",
    "glitch": "glitch.me",
    "cloudfront": "cloudfront.net",
    "fastly": "fastly.net",
    "surge": "surge.sh",
    "h000": "000webhostapp.com",
    "wix": "wixsite.com",
    "weebly": "weebly.com",
    "godaddy": "godaddysites.com",
    "squarespace": "squarespace.com",
    "notion": "notion.site",
    "substack": "substack.com"
}

PLATFORM_CNAME_MARKERS = {
    "vercel": ["vercel-dns.com"],
    "netlify": ["netlify.app", "netlifyglobalcdn.com"],
    "heroku": ["herokudns.com"],
    "google": ["ghs.googlehosted.com", "googlehosted.com"],
    "cloudfront": ["cloudfront.net"],
    "fastly": ["fastly.net"],
    "pages": ["pages.dev"],
    "github": ["github.io"],
    "render": ["render.com"],
    "fly": ["fly.dev"],
    "azurestatic": ["azurestaticapps.net", "azureedge.net"],
    "azureweb": ["azurewebsites.net"],
}

def classify_platform(domain: str, cname_chain: Optional[List[str]] = None) -> Optional[str]:
    dl = domain.lower()
    for k, base in PLATFORM_BASES.items():
        if dl.endswith("." + base) or dl == base or dl.endswith(base):
            return k
    if cname_chain:
        joined = ",".join([c.lower() for c in cname_chain])
        for k, markers in PLATFORM_CNAME_MARKERS.items():
            for m in markers:
                if m in joined:
                    return k
    return None

def get_cname_chain(name: str, timeout: float = 3.0) -> List[str]:
    out = []
    try:
        resolver = dns.resolver.Resolver()
        resolver.lifetime = timeout
        target = name
        seen = set()
        for _ in range(5):
            if target in seen:
                break
            seen.add(target)
            try:
                ans = resolver.resolve(target, "CNAME")
                if not ans:
                    break
                cn = str(ans[0].target).rstrip('.')
                out.append(cn)
                target = cn
            except Exception:
                break
    except Exception:
        return out
    return out

def fetch_ct_keyword(keyword: str) -> List[str]:
    try:
        params = {"q": f"%{keyword}%", "output": "json"}
        r = requests.get("https://crt.sh/", params=params, timeout=20)
        if r.status_code != 200 or not r.text.strip():
            return []
        data = json.loads(r.text)
        out = []
        seen = set()
        for entry in data:
            name = entry.get("name_value")
            if not name:
                continue
            for line in str(name).split("\n"):
                d = line.strip().lower()
                if keyword in d and d not in seen and "." in d:
                    seen.add(d)
                    out.append(d)
        return out
    except Exception:
        return []

def collect_certstream(keywords: List[str], seconds: int = 0, limit: int = 2000) -> List[str]:
    if seconds <= 0:
        return []
    acc = []
    try:
        import certstream
        stop_time = time.time() + seconds
        def callback(message, context):
            if time.time() > stop_time:
                raise SystemExit
            if message.get("message_type") != "certificate_update":
                return
            data = message.get("data", {})
            doms = data.get("leaf_cert", {}).get("all_domains", [])
            for d in doms:
                dl = str(d).lower()
                for kw in keywords:
                    if kw in dl and "." in dl:
                        acc.append(dl)
                        break
            if len(acc) >= limit:
                raise SystemExit
        try:
            certstream.listen_for_events(callback, url='wss://certstream.calidog.io/')
        except SystemExit:
            pass
        except Exception:
            return acc
    except Exception:
        return acc
    return acc

def collect_from_zonefiles(dir_path: str, keywords: List[str], max_lines: int = 0) -> List[str]:
    if not dir_path or not os.path.isdir(dir_path):
        return []
    out = []
    seen = set()
    files = []
    for name in os.listdir(dir_path):
        p = os.path.join(dir_path, name)
        if os.path.isfile(p) and (name.endswith('.txt') or name.endswith('.zone') or name.endswith('.gz')):
            files.append(p)
    for p in files:
        try:
            if p.endswith('.gz'):
                fh = gzip.open(p, 'rt', encoding='utf-8', errors='ignore')
            else:
                fh = open(p, 'r', encoding='utf-8', errors='ignore')
        except Exception:
            continue
        with fh:
            n = 0
            for line in fh:
                if max_lines and n >= max_lines:
                    break
                n += 1
                s = line.strip().lower()
                if not s or s.startswith(';') or ' ' in s:
                    continue
                if '.' not in s:
                    continue
                for kw in keywords:
                    if kw in s:
                        if s not in seen:
                            seen.add(s)
                            out.append(s)
                        break
    return out

def collect_from_sonar(dir_path: str, keywords: List[str], max_lines: int = 0) -> List[Tuple[str, Optional[str]]]:
    if not dir_path or not os.path.isdir(dir_path):
        return []
    out = []
    seen = set()
    files = []
    for name in os.listdir(dir_path):
        p = os.path.join(dir_path, name)
        if os.path.isfile(p) and (name.endswith('.csv') or name.endswith('.csv.gz') or name.endswith('.json') or name.endswith('.json.gz')):
            files.append(p)
    markers = []
    for arr in PLATFORM_CNAME_MARKERS.values():
        markers.extend(arr)
    def parse_line(line: str) -> Tuple[str, str]:
        s = line.strip()
        if not s:
            return '', ''
        if s.startswith('{'):
            try:
                obj = json.loads(s)
                name = str(obj.get('name') or obj.get('domain') or obj.get('host') or '').lower()
                value = str(obj.get('value') or obj.get('data') or obj.get('target') or '').lower()
                return name, value
            except Exception:
                return '', ''
        parts = s.split(',', 1)
        if len(parts) == 1:
            return parts[0].lower(), ''
        return parts[0].lower(), parts[1].lower()
    for p in files:
        try:
            if p.endswith('.gz'):
                fh = gzip.open(p, 'rt', encoding='utf-8', errors='ignore')
            else:
                fh = open(p, 'r', encoding='utf-8', errors='ignore')
        except Exception:
            continue
        with fh:
            n = 0
            for line in fh:
                if max_lines and n >= max_lines:
                    break
                n += 1
                name, value = parse_line(line)
                if not name or '.' not in name:
                    continue
                hit = False
                for kw in keywords:
                    if kw in name:
                        hit = True
                        break
                src = None
                if not hit and value:
                    for m in markers:
                        if m in value:
                            src = classify_platform(name, [value]) or src
                            hit = True
                            break
                if hit:
                    if name not in seen:
                        seen.add(name)
                        out.append((name, src))
    return out

def read_existing_csv(path: str) -> Tuple[List[str], List[Dict[str, str]]]:
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], []
    header = [str(x) for x in rows[0]]
    data = []
    for r in rows[1:]:
        item = {}
        for i, k in enumerate(header):
            item[k] = r[i] if i < len(r) else ""
        data.append(item)
    return header, data

def key_field(header: List[str]) -> str:
    for k in header:
        if str(k).strip().lower() in {"domain", "fqdn"}:
            return k
    return header[0] if header else "domain"

def merge_rows(h1: List[str], rows1: List[Dict[str, str]], h2: List[str], rows2: List[Dict[str, str]]) -> Tuple[List[str], List[Dict[str, str]]]:
    k1 = key_field(h1) if h1 else None
    k2 = key_field(h2) if h2 else None
    header = []
    seenh = set()
    for k in (h1 or []):
        if k not in seenh:
            seenh.add(k)
            header.append(k)
    for k in (h2 or []):
        if k not in seenh:
            seenh.add(k)
            header.append(k)
    if not header and k2:
        header = [k2]
    index = {}
    out = []
    if rows1:
        for r in rows1:
            out.append(dict(r))
            idx_key = r.get(k1) if k1 else None
            if idx_key is not None:
                index[idx_key] = len(out) - 1
    if rows2:
        for r in rows2:
            src_key = r.get(k2) if k2 else None
            if src_key is not None and src_key in index:
                i = index[src_key]
                merged = out[i]
                for k in r.keys():
                    merged[k] = r[k]
                out[i] = merged
            else:
                out.append(dict(r))
                if src_key is not None:
                    index[src_key] = len(out) - 1
    norm_out = []
    for r in out:
        norm = {}
        for k in header:
            norm[k] = r.get(k, "")
        norm_out.append(norm)
    return header, norm_out

def write_csv(path: str, header: List[str], rows: List[Dict[str, str]]):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for r in rows:
            writer.writerow([r.get(k, "") for k in header])
    os.replace(tmp, path)

def process_domain(domain: str, out_dir: str, tlds_path: Optional[str], keywords: List[str], certstream_seconds: int, dict_words: Optional[List[str]], zone_dir: Optional[str], sonar_dir: Optional[str], zone_cap: int, sonar_cap: int) -> Tuple[str, int, int, int]:
    s = time.time()
    h2, r2 = run_dnstwist(domain, tlds_path, dict_words)
    ct_rows = []
    def fetch_platform_kw(kw: str, plat: str, base: str):
        try:
            params = {"q": f"%{kw}%.{base}", "output": "json"}
            r = requests.get("https://crt.sh/", params=params, timeout=20)
            if r.status_code != 200 or not r.text:
                return []
            data = json.loads(r.text)
            out = []
            seen_local = set()
            for entry in data:
                name = entry.get("name_value")
                if not name:
                    continue
                for line in str(name).split("\n"):
                    d = line.strip().lower()
                    if d.endswith(base) and kw in d and d not in seen_local:
                        seen_local.add(d)
                        out.append(d)
            return out
        except Exception:
            return []
    with cf.ThreadPoolExecutor(max_workers=8) as ex:
        futs = []
        for kw in keywords:
            for plat, base in PLATFORM_BASES.items():
                futs.append(ex.submit(fetch_platform_kw, kw, plat, base))
        results = []
        for f in futs:
            try:
                results.extend(f.result())
            except Exception:
                pass
    gen = []
    with cf.ThreadPoolExecutor(max_workers=min(8, max(1, len(keywords)))) as ex:
        futs = [ex.submit(fetch_ct_keyword, kw) for kw in keywords]
        for f in futs:
            try:
                gen.extend(f.result())
            except Exception:
                pass
    if certstream_seconds and certstream_seconds > 0:
        cs = collect_certstream(keywords, seconds=certstream_seconds, limit=4000)
    else:
        cs = []
    z = collect_from_zonefiles(zone_dir, keywords, max_lines=zone_cap) if zone_dir else []
    sonar = collect_from_sonar(sonar_dir, keywords, max_lines=sonar_cap) if sonar_dir else []
    candidates = []
    seen_d = set()
    for d in results + gen + cs + z + [n for n, _ in sonar]:
        if d not in seen_d:
            seen_d.add(d)
            candidates.append(d)
    classified = []
    no_class = []
    for d in candidates:
        src = classify_platform(d, None)
        if src:
            classified.append((d, src))
        else:
            no_class.append(d)
    if no_class:
        with cf.ThreadPoolExecutor(max_workers=16) as ex:
            futs = {ex.submit(get_cname_chain, d): d for d in no_class}
            for f in cf.as_completed(futs):
                d = futs[f]
                try:
                    chain = f.result()
                except Exception:
                    chain = []
                src = classify_platform(d, chain)
                if src:
                    classified.append((d, src))
    for n, src in sonar:
        if src:
            classified.append((n, src))
    # Add CZDS NRD hits and Sonar PDNS hits
    for d in z:
        ct_rows.append({"domain": d, "type": "nrd", "source": "czds"})
    for n, src in sonar:
        ct_rows.append({"domain": n, "type": "pdns", "source": src or "sonar"})
    for d, src in classified:
        ct_rows.append({"domain": d, "type": "ct_log", "source": src})
    if ct_rows:
        if not h2:
            h2 = ["domain", "type", "source"]
        existing_cols = set(h2)
        for col in ["domain", "type", "source"]:
            if col not in existing_cols:
                h2.append(col)
        r2.extend(ct_rows)
    if not h2 and not r2:
        return domain, 0, 0, int(time.time() - s)
    out_path = os.path.join(out_dir, f"{domain}.csv")
    if os.path.isfile(out_path):
        h1, r1 = read_existing_csv(out_path)
        k1 = key_field(h1) if h1 else None
        k2 = key_field(h2) if h2 else None
        existing = set()
        if k1:
            for r in r1:
                v = r.get(k1)
                if v is not None:
                    existing.add(v)
        added = 0
        if k2:
            for r in r2:
                v = r.get(k2)
                if v is not None and v not in existing:
                    added += 1
        h, rows = merge_rows(h1, r1, h2, r2)
        write_csv(out_path, h, rows)
        return domain, added, len(rows), int(time.time() - s)
    else:
        write_csv(out_path, h2, r2)
        return domain, len(r2), len(r2), int(time.time() - s)

def main():
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    xlsx = os.path.join(root, "files", "white_list.xlsx")
    out_dir = os.path.join(root, "sus")
    tlds_path = os.path.join(root, "files", "tlds.txt")
    if not os.path.isfile(tlds_path):
        tlds_path = None
    if not os.path.isfile(xlsx):
        sys.exit(1)
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    kw_map = read_keywords_map(xlsx)
    domains = list(kw_map.keys())
    total = len(domains)
    env_workers = os.getenv("URL_EXTRACTOR_WORKERS")
    try:
        env_workers_val = int(env_workers) if env_workers else None
    except Exception:
        env_workers_val = None
    cpu = os.cpu_count() or 4
    workers = min(total, env_workers_val if env_workers_val and env_workers_val > 0 else cpu)
    cs_secs_env = os.getenv("CERTSTREAM_SECONDS")
    try:
        certstream_seconds = int(cs_secs_env) if cs_secs_env else 0
    except Exception:
        certstream_seconds = 0
    extra_kw_path = os.path.join(root, "files", "keywords.txt")
    extra_kws = read_extra_keywords(extra_kw_path)
    base_extra = [
        "login","secure","verify","update","support","helpdesk","banking","kyc","otp","pay","payment","bill","recharge","claim","policy","ipo","ipoapply","apply","card","credit","debit","netbanking","online","account"
    ]
    seenb = set(extra_kws)
    for w in base_extra:
        if w not in seenb:
            extra_kws.append(w)
            seenb.add(w)
    dict_map: Dict[str, List[str]] = {}
    for d, kws in kw_map.items():
        s = set(kws)
        for w in extra_kws:
            s.add(w)
        dict_map[d] = sorted(list(s))
    zone_dir = os.getenv("CZDS_DIR") or os.path.join(root, "files", "czds")
    if not os.path.isdir(zone_dir):
        zone_dir = None
    sonar_dir = os.getenv("SONAR_DIR") or os.path.join(root, "files", "sonar")
    if not os.path.isdir(sonar_dir):
        sonar_dir = None
    try:
        zone_cap = int(os.getenv("CZDS_MAX_LINES") or 0)
    except Exception:
        zone_cap = 0
    try:
        sonar_cap = int(os.getenv("SONAR_MAX_LINES") or 0)
    except Exception:
        sonar_cap = 0
    print(f"domains:{total} workers:{workers}", flush=True)
    stop = threading.Event()
    start = time.time()
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(process_domain, d, out_dir, tlds_path, kw_map.get(d, []), certstream_seconds, dict_map.get(d), zone_dir, sonar_dir, zone_cap, sonar_cap) for d in domains]
        def status():
            frames = ["|", "/", "-", "\\"]
            i = 0
            while not stop.is_set():
                done = sum(1 for f in futs if f.done())
                running = sum(1 for f in futs if f.running())
                pending = total - done - running
                elapsed = int(time.time() - start)
                s = f"\r[{frames[i]}] running:{running} done:{done}/{total} pending:{pending} elapsed:{elapsed}s"
                print(s, end="", flush=True)
                i = (i + 1) % 4
                time.sleep(0.2)
        th = threading.Thread(target=status, daemon=True)
        th.start()
        done_count = 0
        for f in cf.as_completed(futs):
            try:
                domain, added, total_rows, dur = f.result()
            except Exception:
                domain, added, total_rows, dur = "unknown", 0, 0, 0
            done_count += 1
            sys.stdout.write("\r" + " " * 120 + "\r")
            sys.stdout.flush()
            print(f"[{done_count}/{total}] {domain} added:{added} total:{total_rows} time:{dur}s", flush=True)
        stop.set()
        th.join()
    elapsed = int(time.time() - start)
    print(f"done elapsed:{elapsed}s", flush=True)

if __name__ == "__main__":
    main()